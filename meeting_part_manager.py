"""Contains Meeting Part Manager class"""
import datetime
import itertools
from os import path
from sys import stdin
from time import time
from random import shuffle
from inflect import engine
from openpyxl import Workbook
from openpyxl import load_workbook
from meeting_part import MeetingPart


class MeetingPartManager:
    """Holds and manages multiple objects of the MeetingPart class. Responsible
    for creating a list of objects of the class (one entry for each part),
    randomizing the names returned by the objects without repeating in a column,
    and saving them back to file.

    The class initializes MeetingPart objects for each meeting part, dynamicaly
    calculated by checking the date field. The objects read all the names from
    the corresponding column of the input file, and return them back to this class.
    The names are then stored in 2-D lists (one for friday and one for tuesday).
    Each of these lists is then shuffled ensuring that no name is repeated in any
    given column (thus no one has to do 2 parts on the same day). After getting
    this new list, the names are added back to the MeetingPart class objects.
    A function from those objects is then called to write the now randomized names
    to both sheets of the output file.
    """

    def __init__(self, in_file_name:str = "input_file.xlsx", out_file_name:str = "template.xlsx") -> 'MeetingPartManager':
        """Initialize the Meeting Part Manager class with input file name, output
        file name, and the cell that contains the start date in the input file"""

        input_wb = load_workbook(in_file_name)
        self.input_sheet = input_wb.worksheets[0]
        self.output_wb = load_workbook(out_file_name)
        self.output_sheet = self.output_wb.worksheets[0]
        self.output_sheet2 = self.output_wb.worksheets[1]
        # self.date_cell = date_cell
        self.index_tuesday = {}             # Dict form 'tuesday_names' : 'meeting_part'
        self.index_friday  = {}             # Dict form 'friday_names'  : 'meeting_part'
        self.meeting_parts = []
        self.tuesday_names = []
        self.friday_names  = []

        self.get_start_date()
        self.create_parts()
        # self.shuffle_list(self.tuesday_names)


    def get_start_date(self):
        """Read the start date from the input file. Default to todays date otherwise."""

        self.no_of_parts = 9
        self.start_date = datetime.date.today()
        found_date = False
        second_row = self.input_sheet[2]
        second_row = [second_row[i].value for i in range(len(second_row))]      # Convert row cell to list
        for i in range(len(second_row)):
            if type(second_row[i]) == datetime.datetime:
                self.start_date = second_row[i]
                self.no_of_parts = i - 1            # Do not include the labels column 'A' (-1)
                found_date = True
                break

        if not found_date:
            print("Defaulting to todays date as input for the start date...")


    def create_parts(self):
        """Create objects of class MeetingPart and store them in the meeting_parts list.
        Also add all the names to the names list"""

        for i in range(self.no_of_parts):
            col = chr(ord('A')+i+1)                 # Skip labels column 'A' (+1)
            mp = MeetingPart(self.input_sheet, col)
            if not mp.dead:
                self.meeting_parts.append(mp)
                if mp.tuesday:
                    self.tuesday_names.append(mp.get_names())
                    self.index_tuesday[len(self.tuesday_names)-1] = len(self.meeting_parts)-1
                else:
                    self.friday_names.append(mp.get_names())
                    self.index_friday[len(self.friday_names)-1] = len(self.meeting_parts)-1



    def shuffle_list(self, names: list):
        maxTimeLimit = 0.05
        count = 0
        timeExceded = True
        while timeExceded:
            count += 1
            if count > 100:
                print("\nUnable to find a sutable combination!")
                print("Try and run the program again.")
                print("If the problem persists, there may be a problem with the given input.")
                print("This might be a list of names that make scheduling without repetition on the same day impossible.")
                print("Check input file for potential reasons for this deadlock")
                print("\nExiting program....")
                exit()
            timeExceded = False
            startTime = time()
            shuffle(names[0])
            for i in range(len(names)):
                transposedNames = [list(x) for x in itertools.zip_longest(*names[:i+1])]
                while self.checkDupCols(transposedNames):
                    shuffle(names[i])
                    transposedNames = [list(x) for x in itertools.zip_longest(*names[:i+1])]
                    if( time()-startTime > maxTimeLimit):
                        print(count, "Time taken for randomization was too long")
                        print("Restarting iteration")
                        timeExceded = True
                        break
                if timeExceded:
                    break
        print("Found a combination that has no conflicts")


    def checkDupCols(self, arr):          # Technically it checks duplicate rows, but for a transposed array, so its the original's columns
        for row in arr:
            seen = set()
            for x in row:
                if x in seen:
                    return True
                if x is not None:
                    seen.add(x)
        return False


    def save_to_file(self):
        # Get the next tuesday and friday from the given date (inclusive)
        tuesday = self.start_date + datetime.timedelta((1 - self.start_date.weekday()) % 7)
        friday = self.start_date + datetime.timedelta((4 - self.start_date.weekday()) % 7)
        # Check wether tuesday comes first or not
        tuesday_first = (tuesday < friday)
        if (tuesday_first):
            start_column = 3
        else:
            start_column = 5

        # Save the shuffled list back to the class objects
        for i in range(len(self.tuesday_names)):
            self.meeting_parts[self.index_tuesday[i]].set_shuffled_names(self.tuesday_names[i])
        for i in range(len(self.friday_names)):
            self.meeting_parts[self.index_friday[i]].set_shuffled_names(self.friday_names[i])

        # Write the new list of names to the output file
        for i in range(len(self.meeting_parts)):
            self.meeting_parts[i].write_to_file(self.output_sheet, self.output_sheet2, start_column)

        #Write dates to the top of the file.
        p = engine()
        max_weeks = len( max( max(self.tuesday_names, key=len), max(self.friday_names, key=len), key=len ))
        for i in range(max_weeks):
            date1 = p.ordinal(tuesday.day) + " " + tuesday.strftime('%B')
            date2 = p.ordinal(friday.day) + " " + friday.strftime('%B')
            self.output_sheet.cell(row=2, column=start_column+(i*2)).value = date1
            self.output_sheet.cell(row=2, column=4+(i*2)).value = date2

            tuesday += datetime.timedelta(7)
            friday += datetime.timedelta(7)

        # Save the altered output workbook to a file named Output(n).xlsx with renaming if duplicate present
        output_file_path = "Output"
        index = ''
        while path.isfile(output_file_path+index+".xlsx"):
            if index:
                index = '(' + str(int(index[1:-1]) + 1) + ')'
            else:
                index = '(1)'

        self.output_wb.save(output_file_path+index+".xlsx")
        print("\nOutput file created successfully:", output_file_path+index+".xlsx")
        print("(Press enter to close)")
        stdin.read(1)



if __name__ == '__main__':
    MPM = MeetingPartManager("input_file.xlsx", "template.xlsx")
    MPM.shuffle_list(MPM.tuesday_names)
    MPM.shuffle_list(MPM.friday_names)
    MPM.save_to_file()

    for i in MPM.meeting_parts:
        print(i.part_name)

    MPM.index_tuesday
    MPM.index_friday

    MPM.no_of_parts
    MPM.start_date
    MPM.tuesday_names
    MPM.friday_names
    # MPM.save_to_file()
