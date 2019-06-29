"""Contains Class Meeting Part"""
import datetime
from os import path
from openpyxl import Workbook, load_workbook


class MeetingPart:
    """Holds basic data for a single meeting part,
    and functions to read and write to excel file
    \n
    Variables:
        part_name (str): Name of the Meeting Part
        input_sheet (openpyxl.load_workbook.worksheet): Sheet from the 'input_file' excel document
        column (chr): Column from the input file that contains the relevant data
        tuesday (bool): whether the part is on a Tuesday or not (Friday)
        start_date (datetime.date): the start date for part assignment
        names (list): list of all the names of the people that can be assigned this part
        shuffled_names (list): same list after being randomized

    Methods:
        read_names() : read the names of people from the input excel sheet and
            store them in the 'names' list
        set_shuffled_names(shuffled_names): sets the value for the shuffled names list
        write_to_file (self, output_sheet, output_sheet2, start_column):
            Writes all the shuffled names to the corespoing row/column of both
            sheets of the output file.
    """


    def __init__(self, input_sheet: 'Workbook.worksheets', column: chr) -> 'MeetingPart':
        """Initialize the Meeting Part Class object with apropriate variables"""

        self.input_sheet = input_sheet  # Input worksheet instance from input excel file
        self.column = column            # Column character ('A'/'B'/'C'/...) for reading input file column
        self.names = []
        self.shuffled_names = []
        self.read_input_file()


    def read_input_file(self):
        """Read the names, part name, day and row number from the input sheet and store in the object's variables"""

        # Get the names of the people for the given part
        for cell in self.input_sheet[self.column]:
            if cell.value is not None:
                self.names.append(cell.value)

        self.dead = False
        if(len(self.names)<3):
            self.dead = True
            print("Ignoring column %c from the input file"%self.column)
            return

        try:
            # Get name of the meeting part and its day and output row number
            self.part_name = self.names[0]
            self.tuesday = (self.names[1].lower() == "tuesday")
            self.row = self.names[2]
            if not (type(self.row) == int):
                self.dead = True
                print("Ignoring column %c from the input file"%self.column)
                return
            # Delete the part name, Tuesday/Friday, and row number from the list leaving only peoples names
            self.names.pop(0)
            self.names.pop(0)
            self.names.pop(0)
        except AttributeError:
            self.dead = True
            print("Ignoring column %c from the input file"%self.column)


    def get_names(self):
        """Return the list of names taken from input"""
        return self.names


    def set_shuffled_names(self, shuffled_names: list):
        """Set the value of the list shuffled_names"""
        self.shuffled_names = shuffled_names


    def write_to_file (self, output_sheet, output_sheet2, start_column:int):
        """Write data to output file"""

        # Column index starts from 1, but ('A' - 'A') gives 0 so to compensate (+1)
        # The first column in the input file starts at [column 'B'] = [2], so to go 1 behind the original (-1)
        col = ord(self.column) - ord('A') + 1 - 1       # col -> Column number for output sheet 2 (derived from input column number)

        # Write the shuffled names to both sheets of the output file
        output_sheet2.cell(row=1, column=col).value = self.part_name
        for i in range(len(self.shuffled_names)):
            output_sheet2.cell(row=2+i, column=col).value = self.shuffled_names[i]
            if self.tuesday:
                output_sheet.cell(row=self.row, column=start_column+i*2).value = self.shuffled_names[i]
            else:
                output_sheet.cell(row=self.row, column=4+i*2).value = self.shuffled_names[i]




if __name__ == '__main__':
    from random import shuffle

    input_wb = load_workbook("input_file.xlsx")
    output_wb = load_workbook("template.xlsx")
    input_sheet = input_wb.worksheets[0]
    output_sheet = output_wb.worksheets[0]
    output_sheet2 = output_wb.worksheets[1]

    row = input_sheet[1]
    row = [input_sheet[1][i].value for i in range(len(input_sheet[1]))]
    row
    row.index("Enter the start date below") + 1

    found_date = False
    second_row = input_sheet[2]
    second_row = [second_row[i].value for i in range(len(second_row))]      # convert row cell to row list
    for i in range(len(second_row)):
        x = second_row[i]
        if type(second_row[i]) == datetime.datetime:
            start_date = second_row[i]
            no_of_parts = i - 1
            found_date = True
            break
    start_date
    no_of_parts
    found_date
    type(second_row[12]) == datetime.datetime
    second_row


    col = chr(ord('A')+1)
    # input_sheet[self.column]
    chairman = MeetingPart(input_sheet, col)
    chairman.part_name
    chairman.tuesday
    chairman.row
    chairman.names
    chairman.get_names()
    shuffle(chairman.names)
    chairman.set_shuffled_names(chairman.names)
    chairman.shuffled_names
    chairman.write_to_file(output_sheet, output_sheet2, 3)

    for i in range(1, 9):
        col = chr(ord('A') + i)
        mp = MeetingPart(input_sheet, col)
        print(mp.names)
        print(mp.part_name)
        print(mp.get_names())

    output_file_path = "Output"
    index = ''
    while path.isfile(output_file_path+index+".xlsx"):
        if index:
            index = '(' + str(int(index[1:-1]) + 1) + ')'
        else:
            index = '(1)'
    index
    output_wb.save(output_file_path+index+".xlsx")
    print("Output file created successfully:", output_file_path+index+".xlsx")
