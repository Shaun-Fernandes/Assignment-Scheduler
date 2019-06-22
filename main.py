import datetime
from os import path
from sys import stdin
from random import shuffle
from inflect import engine
from openpyxl import Workbook, load_workbook


def main():
    check_files()
    input_wb = load_workbook("input_file.xlsx")
    output_wb = load_workbook("template.xlsx")
    input_sheet = input_wb.worksheets[0]
    output_sheet = output_wb.worksheets[0]
    output_sheet2 = output_wb.worksheets[1]
    p = engine()
    wt_readers = []
    cbs_readers = []

###########################
# Put all this into a class
    # Retrive all names in column 1 (A)
    for cell in input_sheet['A']:
        if cell.value is not None:
            wt_readers.append(cell.value)

    # Retrive all names in column 2 (B)
    for cell in input_sheet['B']:
        if cell.value is not None:
            cbs_readers.append(cell.value)

    # Delete the column headders
    wt_readers.pop(0)
    cbs_readers.pop(0)

    # Randomize both lists
    shuffle(wt_readers)
    shuffle(cbs_readers)
# Tll here
##########

#############################
# Make all of this a function
    # Get the start date from the input file
    entered_date = input_sheet['C2'].value
    if (entered_date is not None) and not (type(entered_date) == datetime.datetime):
        print("Unable to read the entered date.")
        print("In the input file, please change the format of the cell for the start date to type 'date' (short date or long date will do)\n")
    if (entered_date is not None) and (type(entered_date) == datetime.datetime):
        start_date = entered_date
    else:
        print("Defaulting to todays date as input for the start date...")
        start_date = datetime.date.today()
    start_date
    # Get the next tuesday and friday from the given date (inclusive)
    tuesday = start_date + datetime.timedelta((1 - start_date.weekday()) % 7)
    friday = start_date + datetime.timedelta((4 - start_date.weekday()) % 7)

    tuesday
    friday

    # Check wether tuesday comes first or not
    # tuesday_first = (tuesday < friday)
    # start = 3 if(tuesday_first) else 5
    if (tuesday < friday):
        tuesday_first = True
        start = 3
    else:
        tuesday_first = False
        start = 5
# Till here
#############################

################################
# Put all of this into the class
    # Write the list of CBS readers
    for i in range(len(cbs_readers)):
        output_sheet.cell(row=13, column=start+i*2).value = cbs_readers[i]
        output_sheet2.cell(row=2+i, column=2).value = cbs_readers[i]

    # Write the list of WT readers
    for i in range(len(wt_readers)):
        output_sheet.cell(row=13, column=4+i*2).value = wt_readers[i]
        output_sheet2.cell(row=2+i, column=1).value = wt_readers[i]
# Till here
###########

#############################
# Make all of this a function
    # Write the dates to the top of the schedule
    max_weeks = max(len(wt_readers), len(cbs_readers))
    max_weeks
    for i in range(max_weeks):
        date1 = p.ordinal(tuesday.day) + " " + tuesday.strftime('%B')
        date2 = p.ordinal(friday.day) + " " + friday.strftime('%B')
        output_sheet.cell(row=2, column=start+(i*2)).value = date1
        output_sheet.cell(row=2, column=4+(i*2)).value = date2

        tuesday += datetime.timedelta(7)
        friday += datetime.timedelta(7)
# Till here
#############################

    # Deleting unnecessary columns (Carzy math xD. BA=53, 'DA' is arbitary)
    output_sheet.delete_cols(start+(max_weeks-1)*2+1, amount=30)
    # output_sheet.move_range("BA1:DA14", cols=((start+(max_weeks-1)*2+1)-53))
    # if not tuesday_first:
    #     # output_sheet.delete_cols(3,1)
    #     output_sheet.move_range("D1:BA14", cols=-1)

#############################
# Make all of this a function
    # Save the whole thing to a file named output with renaming if needed
    output_file_path = "Output"
    index = ''
    while path.isfile(output_file_path+index+".xlsx"):
        if index:
            index = '(' + str(int(index[1:-1]) + 1) + ')'
        else:
            index = '(1)'

    output_wb.save(output_file_path+index+".xlsx")
    print("Output file created successfully:", output_file_path+index+".xlsx")
    print("(Press enter to close)")
    stdin.read(1)
# Till here
#############################


def check_files():
    if not path.isfile("template.xlsx") and not path.isfile("input_file.xlsx"):
        print("Please add a file called 'input_file.xlsx' and 'template.xlsx' in this folder")
        print("(Press enter to continue)")
        stdin.read(1)
        exit()
    if not path.isfile("input_file.xlsx"):
        print("Please add a file called 'input_file.xlsx' in this folder")
        print("(Press enter to continue)")
        stdin.read(1)
        exit()
    if not path.isfile("template.xlsx"):
        print("Please add a file called 'template.xlsx' in this folder")
        print("(Press enter to continue)")
        stdin.read(1)
        exit()


if __name__ == '__main__':
    main()
